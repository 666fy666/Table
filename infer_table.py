import logging
import os
import time

import numpy as np
import openpyxl
from PIL import Image
from bs4 import BeautifulSoup
from lineless_table_rec import LinelessTableRecognition
from lineless_table_rec.utils_table_recover import format_html, plot_rec_box_with_logic_info, plot_rec_box
from openpyxl.styles import Alignment, Side, Border
from openpyxl.utils import get_column_letter
from rapid_undistorted.inference import InferenceEngine
from rapidocr_paddle import RapidOCR
from table_cls import TableCls
from wired_table_rec import WiredTableRecognition

# 配置日志（确保只调用一次）
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')


def rec_table(img_path, output, pre=False, logic_info=False, box=False):
    start = time.perf_counter()
    image = Image.open(img_path)
    img = np.array(image)

    img = check_rotate(img)  # 检测图片旋转角

    name = img_path.split('/')[-1].split('.')[0]

    if pre:
        # 扭曲修正->去阴影->去模糊 (指定去模糊模型)
        engine = InferenceEngine()
        img, elapse = engine(img, ["unwrap", "unshadow", ("unblur", "OpenCvBilateral")])

    # 默认小yolo模型(0.1s)，可切换为精度更高yolox(0.25s),更快的qanything(0.07s)模型
    table_cls = TableCls(model_type="q")  # TableCls(model_type="yolox"),TableCls(model_type="q")
    cls, elasp = table_cls(img)

    lineless_engine = LinelessTableRecognition()
    wired_engine = WiredTableRecognition()
    if cls == 'wired':
        table_engine = wired_engine
    else:
        table_engine = lineless_engine

    # html, elasp, polygons, logic_points, ocr_res = table_engine(output_img)
    # logging.info(f"elasp: {elasp}")

    # 使用其他ocr模型,（onnx或者paddle的infer）

    ocr_engine = RapidOCR(det_model_dir="./Models/det_infer.onnx",
                          rec_model_dir="./Models/rec_infer.onnx",
                          det_use_cuda=True,
                          cls_use_cuda=True,
                          rec_use_cuda=True,
                          use_det=True,
                          use_cls=True,
                          use_rec=True
                          )
    ocr_res, _ = ocr_engine(img)

    # print(ocr_res)

    # 表格识别模型
    html, elasp, polygons, logic_points, ocr_res = table_engine(
        img,
        ocr_result=ocr_res,
        col_threshold=50,  # 识别框左边界x坐标差值小于col_threshold的默认同列
        row_threshold=10,  # 识别框上边界y坐标差值小于row_threshold的默认同行
        rec_again=False
    )
    complete_html = format_html(html)

    # 确保输出文件夹存在
    if not os.path.exists(output):
        os.makedirs(output)

    '''
    os.makedirs(os.path.dirname(f"{output}/{name}.html"), exist_ok=True)
    with open(f"{output}_output/{name}.html", "w", encoding="utf-8") as file:
        file.write(complete_html)
    '''

    # html转excel
    html_table_to_excel(complete_html, f"{output}/{name}.xlsx")

    if logic_info:
        # 可视化表格识别框 + 逻辑行列信息
        plot_rec_box_with_logic_info(
            img_path, f"{output}/{name}_rec_box.jpg", logic_points, polygons
        )

    if box:
        # # 可视化 ocr 识别框
        plot_rec_box(img_path, f"{output}/{name}_ocr_box.jpg", ocr_res)

    end = time.perf_counter()
    elapsed_time = end - start
    print(f"Finished:{img_path},Time:{elapsed_time}")
    logging.info(f"Finished:{img_path},Time:{elapsed_time}")





def html_table_to_excel(html, excel_path):
    # 解析HTML
    soup = BeautifulSoup(html, 'html.parser')
    table = soup.find('table')
    if table is None:
        return

    # 初始化Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # 创建黑色边框样式
    black_side = Side(border_style='thin', color='000000')  # 新增边框样式
    black_border = Border(left=black_side, right=black_side,
                         top=black_side, bottom=black_side)

    # 计算最大列数
    rows = table.find_all('tr')
    max_cols = 0
    for row in rows:
        col_counter = 0
        for cell in row.find_all(['td', 'th']):
            colspan = int(cell.get('colspan', 1))
            col_counter += colspan
        max_cols = max(max_cols, col_counter)

    # 创建动态扩展矩阵
    matrix = []
    for _ in range(len(rows)):
        matrix.append([None] * max_cols)

    # 存储合并信息
    merge_list = []

    # 填充矩阵
    for row_idx, row in enumerate(rows):
        col_idx = 0
        cells = row.find_all(['td', 'th'])

        for cell in cells:
            # 跳过已填充位置
            while col_idx < max_cols and matrix[row_idx][col_idx] is not None:
                col_idx += 1

            if col_idx >= max_cols:
                break

            # 获取单元格属性
            rowspan = int(cell.get('rowspan', 1))
            colspan = int(cell.get('colspan', 1))
            text = cell.get_text(separator='\n').strip()

            # 动态扩展行
            while row_idx + rowspan > len(matrix):
                matrix.append([None] * max_cols)

            # 记录合并信息
            merge_list.append((
                row_idx + 1,  # Excel行号从1开始
                col_idx + 1,  # Excel列号从1开始
                row_idx + rowspan,
                col_idx + colspan
            ))

            # 填充主单元格
            matrix[row_idx][col_idx] = text

            # 标记被合并的单元格
            for r in range(rowspan):
                for c in range(colspan):
                    if r == 0 and c == 0:
                        continue  # 跳过主单元格
                    if (row_idx + r) < len(matrix) and (col_idx + c) < max_cols:
                        matrix[row_idx + r][col_idx + c] = ""

            col_idx += colspan

    # 写入Excel（添加边框设置）
    alignment = Alignment(wrap_text=True, vertical='top', horizontal='center')

    for row_idx, row_data in enumerate(matrix):
        for col_idx, cell_value in enumerate(row_data):
            if cell_value is not None:
                # 获取单元格对象并设置样式
                cell = ws.cell(row=row_idx + 1, column=col_idx + 1, value=cell_value)
                cell.alignment = alignment
                cell.border = black_border  # 应用黑色边框

    # 应用合并
    for merge in merge_list:
        start_row, start_col, end_row, end_col = merge
        if start_row != end_row or start_col != end_col:
            ws.merge_cells(start_row=start_row, start_column=start_col,
                           end_row=end_row, end_column=end_col)

    # 自动调整列宽（优化版）
    for col_idx in range(1, max_cols + 1):
        max_length = 0
        column_letter = get_column_letter(col_idx)
        for cell in ws[column_letter]:
            if cell.value:
                cell_lines = str(cell.value).split('\n')
                max_line_length = max(len(line) for line in cell_lines)
                if max_line_length > max_length:
                    max_length = max_line_length
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = adjusted_width

    # 保存文件
    wb.save(excel_path)

def check_rotate(image):

    return image



def subfolder_dir(path):
    # 遍历主文件夹下的所有子文件夹
    for entry in os.scandir(path):
        if entry.is_dir():  # 检查是否是子文件夹
            subfolder_path = entry.path
            one_dir(path=subfolder_path, output=subfolder_path)


def one_dir(path,output):
    # 遍历文件夹中的所有文件
    for filename in os.listdir(path):
        # 检查文件是否为图片（可以根据需要添加更多图片格式）
        if filename.lower().endswith(('.png', '.jpg', '.jpeg', '.bmp', '.gif', '.tiff')):
            img_path = os.path.join(path, filename)
            # 调用 rec_table 函数处理图片
            rec_table(img_path=img_path,output=output)


if __name__ == '__main__':
    # 测试单张图片
    rec_table(img_path="./79c35a5dc9a49bf525942f81eef3087e_compress.jpg",output="./")
    rec_table(img_path="./pic(5)(1).jpg", output="./")
    rec_table(img_path="./k1.jpg", output="./")

    # 测试单个文件夹（文件夹里全是图片）
    #one_dir(path = "./test",output="./test_output")

    # 测试子文件夹文件夹（大文件夹里全是子文件夹，图片在在文件夹里,结果也保存到子文件夹）
    #subfolder_dir(path = "./copied_folders")






